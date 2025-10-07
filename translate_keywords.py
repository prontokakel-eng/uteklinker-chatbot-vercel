import os
import json
import math
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.worksheet.datavalidation import DataValidation

from dotenv import load_dotenv
from openai import OpenAI

# Ladda env.local från projektroten
load_dotenv(dotenv_path=".env.local", override=True)


# Initiera klienten (hämtar API-nyckeln från env.local)
client = OpenAI()

def chat_completion(messages, model="gpt-4o", temperature=0):
    resp = client.chat.completions.create(
        model=model,
        messages=messages,
        temperature=temperature
    )
    return resp.choices[0].message.content


# ======== CONFIG ========
INPUT_XLSX = r"C:\Users\chris\Documents\SE_FULL_LOOKUP.xlsx"
SHEET_SE_LOOKUP = "SE_FULL_LOOKUP"
OUTPUT_XLSX = r"C:\Users\chris\Documents\Faq_keywords_lookup_multilang.xlsx"
OUTPUT_JSON = r"C:\Users\chris\Documents\Faq_keywords_lookup_multilang.json"

MODEL_TRANSLATE = "gpt-4o"     # or "gpt-4o" for best quality
MODEL_SYNONYMS  = "gpt-4o"     # can be same as above
BATCH_SIZE = 120                    # safe batch size

# ======== helpers ========
def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i+n]

def translate_batch(sw_words, target_lang):
    prompt = (
        f"Translate the following Swedish keywords into {target_lang}.\n"
        f"Return ONLY a plain list, one translation per line, same order, no numbering.\n\n"
        "Swedish words:\n" + "\n".join(sw_words)
    )
    txt = chat_completion(
        [{"role": "user", "content": prompt}],
        model=MODEL_TRANSLATE,
        temperature=0
    )
    lines = [x.strip() for x in txt.splitlines() if x.strip()]
    # If model returns fewer lines than input, pad with originals to keep length
    if len(lines) != len(sw_words):
        # attempt a best-effort map; if mismatch, fill with Swedish word
        lines = (lines + sw_words)[:len(sw_words)]
    return lines

def generate_synonyms_batch(words, lang_label):
    # Ask for 1–3 simple synonyms per word, comma-separated, or "-" if none.
    prompt = (
        f"Generate up to 3 common synonyms for each of the following {lang_label} keywords.\n"
        f"Return ONLY a plain list, one line per word, comma-separated synonyms, or '-' if none.\n\n"
        f"Words:\n" + "\n".join(words)
    )
    txt = chat_completion(
        [{"role": "user", "content": prompt}],
        model=MODEL_SYNONYMS,
        temperature=0
    )
    lines = [x.strip() for x in txt.splitlines()]
    # length guard
    if len(lines) != len(words):
        lines = (lines + ["-"] * len(words))[:len(words)]
    return lines

def categorize_se(se_word: str) -> str:
    w = (se_word or "").lower()
    if any(x in w for x in ["leverans", "hemleverans", "frakt", "pris", "avgift", "kostnad", "tid", "leveranstid"]):
        return "Leverans"
    if any(x in w for x in ["olja", "underhåll", "underhållsolja", "skötsel", "impregnering", "mossa", "alger", "salt", "kemikalier", "frost", "vinter", "rengöring", "tvätt"]):
        return "Underhåll"
    if any(x in w for x in ["klinker", "klinkerdäck", "platta", "plattor", "betong", "mått", "storlek", "dimension", "ytstruktur", "struktur", "färg", "design", "vikt", "tjocklek", "kapning", "sågning"]):
        return "Produkt"
    return "Övrigt"

def is_domain_term(se_word: str) -> bool:
    w = (se_word or "").lower()
    domain_terms = [
        "klinker","klinkerdäck","betong","impregnering","underhåll","underhållsolja",
        "skötsel","mossa","alger","frost","salt","kemikalier","ytstruktur","kapning",
        "sågning","hemleverans","leverans"
    ]
    return any(x in w for x in domain_terms)

def make_check(se, en, da, de):
    se_ = (se or "").strip().lower()
    return "NEED_CHECK" if (se_.lower() in [(en or "").strip().lower(),
                                            (da or "").strip().lower(),
                                            (de or "").strip().lower()]) else "OK"

# ======== main pipeline ========
def main():
    # 1) Load Swedish keywords
    df_se = pd.read_excel(INPUT_XLSX, sheet_name=SHEET_SE_LOOKUP)
    if "keyword" not in df_se.columns:
        raise ValueError(f"Sheet '{SHEET_SE_LOOKUP}' must contain a 'keyword' column.")
    se_words = df_se["keyword"].dropna().astype(str).tolist()
    print(f"Loaded {len(se_words)} Swedish keywords.")

    # 2) Translate to EN / DA / DE in batches
    en_all, da_all, de_all = [], [], []
    for batch in chunks(se_words, BATCH_SIZE):
        en_all += translate_batch(batch, "English")
        da_all += translate_batch(batch, "Danish")
        de_all += translate_batch(batch, "German")

    # 3) Synonyms per language (also batched)
    syn_en, syn_da, syn_de = [], [], []
    for batch_en in chunks(en_all, BATCH_SIZE):
        syn_en += generate_synonyms_batch(batch_en, "English")
    for batch_da in chunks(da_all, BATCH_SIZE):
        syn_da += generate_synonyms_batch(batch_da, "Danish")
    for batch_de in chunks(de_all, BATCH_SIZE):
        syn_de += generate_synonyms_batch(batch_de, "German")

    # 4) Assemble DataFrame with extra columns
    data = {
        "SE": se_words,
        "EN": en_all,
        "DA": da_all,
        "DE": de_all,
        "SYNONYMS_EN": syn_en,
        "SYNONYMS_DA": syn_da,
        "SYNONYMS_DE": syn_de,
    }
    df = pd.DataFrame(data)
    df["CHECK"] = [make_check(se, en, da, de) for se, en, da, de in zip(df.SE, df.EN, df.DA, df.DE)]
    df["CATEGORY"] = df["SE"].apply(categorize_se)
    df["DOMAIN"] = df["SE"].apply(lambda x: "TRUE" if is_domain_term(x) else "FALSE")
    df["NOTES"] = ""

    # 5) Save Excel (single sheet LOOKUP_ALL)
    with pd.ExcelWriter(OUTPUT_XLSX, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="LOOKUP_ALL", index=False)

    # 6) Add conditional formatting & validation
    wb = load_workbook(OUTPUT_XLSX)
    ws = wb["LOOKUP_ALL"]

    # Find column indexes
    headers = {cell.value: i+1 for i, cell in enumerate(ws[1])}
    col_CHECK = headers["CHECK"]
    col_CATEGORY = headers["CATEGORY"]

    # Conditional fill styles
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # light red
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # light green
    blue_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")    # light blue (Produkt)
    yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow (Leverans)
    light_green = PatternFill(start_color="E2F0D9", end_color="E2F0D9", fill_type="solid")  # light green (Underhåll)

    last_row = ws.max_row

    # CHECK rules
    dxf_red = DifferentialStyle(fill=red_fill)
    dxf_green = DifferentialStyle(fill=green_fill)
    ws.conditional_formatting.add(f"{ws.cell(row=2, column=col_CHECK).coordinate}:{ws.cell(row=last_row, column=col_CHECK).coordinate}",
                                  CellIsRule(operator="equal", formula=['"NEED_CHECK"'], dxf=dxf_red))
    ws.conditional_formatting.add(f"{ws.cell(row=2, column=col_CHECK).coordinate}:{ws.cell(row=last_row, column=col_CHECK).coordinate}",
                                  CellIsRule(operator="equal", formula=['"OK"'], dxf=dxf_green))

    # CATEGORY color bands (simple per-cell coloring)
    for row in range(2, last_row+1):
        cat = ws.cell(row=row, column=col_CATEGORY).value or ""
        if cat == "Produkt":
            ws.cell(row=row, column=col_CATEGORY).fill = blue_fill
        elif cat == "Leverans":
            ws.cell(row=row, column=col_CATEGORY).fill = yellow_fill
        elif cat == "Underhåll":
            ws.cell(row=row, column=col_CATEGORY).fill = light_green

    # Optional: Data validation dropdown for CATEGORY (helps manual editing)
    dv = DataValidation(type="list", formula1='"Produkt,Leverans,Underhåll,Övrigt"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{ws.cell(row=2, column=col_CATEGORY).coordinate}:{ws.cell(row=last_row, column=col_CATEGORY).coordinate}")

    wb.save(OUTPUT_XLSX)

    # 7) Save JSON (arrays per language)
    out_json = {
        "SE": df["SE"].tolist(),
        "EN": df["EN"].tolist(),
        "DA": df["DA"].tolist(),
        "DE": df["DE"].tolist(),
    }
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out_json, f, ensure_ascii=False, indent=2)

    print("✅ Done!")
    print(f"Excel: {OUTPUT_XLSX}")
    print(f"JSON : {OUTPUT_JSON}")

if __name__ == "__main__":
    main()
